import openpyxl, os, argparse, re, sys
from typing import Tuple
from openpyxl.utils import column_index_from_string, get_column_letter, exceptions
from openpyxl.styles import Font


def safe_open_and_close(file_path: str) -> None:
    """
    安全地打开并 Excel工作簿，打印操作结果

    :param file_path: 文件路径
    :return: 返回为None
    """
    wb = None

    try:
        # 检验是否存在
        if not os.path.exists(file_path):
            print("错误：文件不存在")
            return

        # 检查是否可读
        if not os.access(file_path, os.R_OK):
            print("错误：文件不可读")
            sys.exit()
            return

        wb = openpyxl.load_workbook(file_path)
        sheet_count = len(wb.sheetnames)
        active_sheet = wb.active.title
        wb.save(file_path)
        print("成功打开工作簿")

    except exceptions.InvalidFileException:
        print("错误：无效的Excel文件")
        sys.exit()

    except PermissionError:
        print("错误：文件被拒绝访问")
        sys.exit()

    except Exception as e:
        print("打开文件时发生未知错误")
        sys.exit()

    finally:
        if wb:
            wb.close()
            print("工作簿已安全关闭")
        else:
            print("未打开工作簿，无需关闭")	


def get_workbook(subject: str, path: str) -> str:
    """
    获取工作簿函数
    
    :param subject: 工作簿名称中含有的科目名称
    :param path: 工作簿目录
    :return: 指定的工作簿名称
    
    """
    all_files = os.listdir(path)
    os.chdir(path)
    excel_files = [file_name for file_name in all_files if file_name.endswith('.xlsx')]
    selected_workbook = [test for test in excel_files if subject in test]
    selected_workbook_num = len(selected_workbook)
    real_selected_workbook = ""
    
    if selected_workbook_num == 0 and len(excel_files) > 0:
        print(f"您所在的目录似乎没有{subject}类成绩相关文件\n但是您可以选择打开指定的文件以进入指定工作簿")

        while True:
            try:
                selected_file_input = int(input("请您输入您想要进行操作的工作表："))
            except:
                print("请不要输入非数字")
            else:

                if selected_workbook_num == "exit":
                    break
                elif selected_file_input > selected_workbook_num or selected_file_input < 0:
                    print("请不要输入非显示序号的数字！")
                else:
                    print("已完成输入！")
                    break

    elif selected_workbook_num == 0 and len(excel_files) == 0:
        print(f"这里没有{subject}类工作簿")
    elif selected_workbook_num == 1:
        real_selected_workbook = selected_workbook[selected_workbook_num - 1]
        
    else:
        print(f"您所在的目录存在多个含有{subject}类的工作簿\n您可以选择进入指定工作簿")
        try:
            for test in range(int(selected_workbook_num)):
                print(f"{test+1}.{selected_workbook[test]}")
        except:
            print("发生错误！")
        
        #  循环以进行询问    
        while True:
            try:
                selected_workbook_num_two = int(input("请您输入您想要进行操作的工作簿："))
            except:
                print("请不要输入非数字")
            else:

                if selected_workbook_num_two == "exit":
                    break

                elif selected_workbook_num_two > selected_workbook_num or selected_workbook_num_two <= 0:
                    print("请不要输入非显示序号的数字！")

                else:
                    print("已完成输入！")
                    real_selected_workbook = selected_workbook[selected_workbook_num_two - 1]
                    break

    safe_open_and_close(path + "\\" + real_selected_workbook)

    return real_selected_workbook


def get_sheet(workbook: str, sheet: str, subject: str) -> str:
    """
    获取工作簿中指定的工作表

    :param workbook: 工作表所在工作簿（.xlsx文件）
    :param sheet: 工作表名称
    :param subject: 工作表名称中所含学科关键词
    :return: 工作表名称
    """
    sheet_name = ''

    if sheet is not None:
        sheet_name = sheet

    else:
        wb = openpyxl.load_workbook(workbook)
        sheet_num = len(wb.sheetnames)

        sheet_sub = [test for test in wb.sheetnames if subject in test]
        sheet_sub_num = len(sheet_sub)
        
        if sheet_num == 1 and sheet_sub_num == 1:
            sheet_name = wb.sheetnames[0]
        elif sheet_num != 1 and sheet_sub_num == 1:
            sheet_name = sheet_sub[0]
        elif sheet_num == 0 or sheet_sub_num == 0:
            print(f"您选择的工作簿没有{subject}类工作表")
        else:
            
            print(f"您选择的工作簿中有多个{subject}类工作表\n您可以进行选择")
            
            for test in range(len(sheet_sub)):
                print(f"{test+1}.{sheet_sub[test]}")
                
            while True:
                try:
                    selected_sheet_num = int(input("请您输入您想要进行操作的工作表："))
                except:
                    print("请不要输入非数字")
                else:
                    if selected_sheet_num > sheet_sub_num or selected_sheet_num < 0:
                        print("请不要输入非显示序号的数字！")

                    else:
                        print("已完成输入！")
                        sheet_name = wb.sheetnames[selected_sheet_num - 1]
                        break
                        
        wb.close()
    
    return sheet_name


def get_data_place(workbook: str, sheet: str) -> Tuple[int, int, str]:
    """
    获取工作表内容第一次出现的位置

    :param workbook:  指定工作簿（后缀为.xlsx的文件）
    :param sheet:   指定工作表
    :return: 工作表最大行， 工作表最大列， 工作表内容开始位置
    """
    wb = openpyxl.load_workbook(workbook)
    new_sheet = wb[sheet]

    wb_row_num = new_sheet.max_row
    wb_column_num = new_sheet.max_column
    result_cell = ''

    if new_sheet['A1'].value is None:
        if new_sheet['A2'].value is None or new_sheet['A2'].value == " ":
            print("A1和A2都为空！")
            if new_sheet['B1'].value is None or new_sheet['B1'].value == " ":
                print("A1和B1都为空！")

            else:
                print(f"{workbook}中\n{sheet}A1无内容, B1有内容")
                result_cell = "B1"

        else:
            print(f"{workbook}中\n{sheet}A1无内容，A2有内容")
            result_cell = "A2"

    else:
        print(f"{workbook}中\n{sheet}A1有内容")
        result_cell = "A1"

    wb.close()

    return wb_row_num, wb_column_num, result_cell


def verify_title(workbook: str, sheet: str, start_cell: str) -> Tuple[int, int, str, bool]:
    """
    确定工作表有无标题，标题大小及内容

    :param workbook: 指定工作簿（后缀为.xlsx的文件）
    :param sheet: 指定工作表
    :param start_cell: 工作表内容起始位置
    :return: 标题行大小，标题列大小，标题值，是否存在标题
    """
    wb = openpyxl.load_workbook(workbook)
    sheet = wb[sheet]

    # 对变量进行初赋值
    row_size = 0
    column_size = 0
    default_value = ""
    exist = False

    # 是否存在标题等系列操作
    merged_ranges = sheet.merged_cells.ranges
    for merged_cell in merged_ranges:
        if start_cell in merged_cell.coord:
            exist = True
            row_size = merged_cell.max_row - merged_cell.min_row + 1
            column_size = merged_cell.max_col - merged_cell.min_col + 1
            default_value = sheet[start_cell].value

    wb.close()

    if exist:
        print(f"工作表存在标题，标题是{default_value}\n标题占{row_size}行，{column_size}列")
    else:
        print("工作表似乎不存在标题")

    return row_size, column_size, default_value, exist


def get_sub_title(workbook: str, sheet: str, start_cell: str, title_row_size: int, title_col_size: int, max_column: int) -> Tuple[str, dict]:
    """
    获取表头信息

    :param workbook: 指定工作簿（后缀为.xlsx的文件）
    :param sheet: 指定工作表
    :param start_cell: 工作表内容起始位置
    :param title_row_size: 标题的行大小
    :param title_col_size: 标题的列大小
    :param max_column: 工作表最大列
    :return: 表头起始位置，表头位置及至对应的字典
    """
    wb = openpyxl.load_workbook(workbook)
    sheet = wb[sheet]

    start_sub_title_row = 1 + title_row_size
    start_sub_title_col = re.match(r"^([A-Za-z]+)", start_cell).group(1).upper()
    start_sub_title = start_sub_title_col + str(start_sub_title_row)
    sub_title_dict = {}

    # 获取字典
    for test in range(max_column - column_index_from_string(start_sub_title_col) + 1):
        place = get_column_letter(column_index_from_string(start_sub_title_col) + test) + str(start_sub_title_row)
        sub_title_dict[place] = sheet[place].value

    wb.close()

    print(f"表头位置及内容的字典：\n{sub_title_dict}")

    return start_sub_title, sub_title_dict


def verify_heading_three(workbook: str, sheet: str, start_sub_title: str, max_column: int) -> Tuple[list, bool]:
    """
    确认工作表有无表头，表头内容

    :param workbook: 指定工作簿（后缀为.xlsx的文件）
    :param sheet: 指定工作表
    :param start_sub_title: 工作表表头起始位置
    :param max_column: 工作表最大列
    :return: 如果次表头存在的表次头名称列表，是否存在次标头
    """
    wb = openpyxl.load_workbook(workbook)
    sheet = wb[sheet]

    # 官方文件必有表头，故以表头为基础，次表头与之对应，只存入列表
    heading_three = []
    # 注意，这里的row已转化为整型
    sub_title_row = int(''.join(re.findall(r'\d+', start_sub_title)))
    sub_title_col = re.match(r"^([A-Za-z]+)", start_sub_title).group(1).upper()
    exist = False

    # 判断指定的表头下的行的最后一列是否为数字，如果是则证明无次表头，如果是字符串，则有次表头
    if isinstance(sheet.cell(row=sub_title_row + 1,
                             column=max_column).value, int):
        heading_three = []

    elif isinstance(sheet.cell(row=sub_title_row + 1,
                               column=max_column).value, str):

        for test in range(column_index_from_string(sub_title_col), max_column + 1):
            heading_three.append(sheet.cell(row=sub_title_row + 1, column=test).value)
            exist = True

    wb.close()

    if exist:
        print(f"工作表存在次表头次表头列表是\n{heading_three}")
    else:
        print("工作表不存在次表头")

    return heading_three, exist


def verify_school(workbook: str, sheet: str, max_row: int, sub_title_dict: dict, title_row_size: int, heading_three_exist: bool, school: str) -> Tuple[str, dict]:
    """
    筛选出指定学校的学生

    :param workbook: 指定工作簿（后缀为.xlsx的文件）
    :param sheet: 指定工作表
    :param max_row: 工作表最大行
    :param sub_title_dict: 表头位置及值对应的字典
    :param title_row_size: 标题行大小
    :param heading_three_exist: 次标头是否存在
    :param school: 指定筛选的学生所在学校
    :return: school在sub_title的位置，属于school的学生姓名出现的位置及学生姓名的字典
    """
    wb = openpyxl.load_workbook(workbook)
    sheet = wb[sheet]

    in_school_dict = {}
    school_cell = ""
    # 在表头中获取指定文字位置

    try:
        school_cell = [key for key, value in sub_title_dict.items() if value == "学校" or value == "学校名称"][0]
        school_cell_row = ''.join(re.findall(r'\d+', school_cell))
        school_cell_col = re.match(r"^([A-Za-z]+)", school_cell).group(1).upper()

        if heading_three_exist:
            start_row = int(school_cell_row) + 1 + title_row_size
        else:
            start_row = int(school_cell_row) + title_row_size

    except IndexError:
        print("表头没有\"学校\"或\"学校名称\"")

    else:
        try:
            name_cell = [key for key, value in sub_title_dict.items() if value == "姓名"][0]
            name_cell_col = re.match(r"^([A-Za-z]+)", name_cell).group(1).upper()

            for row in range(start_row, max_row + 1):
                school_place = f"{school_cell_col}{row}"
                name_place = f"{name_cell_col}{row}"
                if school is None:
                    in_school_dict[name_place] = sheet[name_place].value
                else:
                    if school in sheet[school_place].value:
                        in_school_dict[name_place] = sheet[name_place].value

            # print(f"in_school_dict的内容是\n{in_school_dict}")

        except IndexError:
            print("表头没有\"姓名\"")

    wb.close()

    return school_cell, in_school_dict


def verify_class(workbook: str, sheet: str, max_row: int, sub_title_dict: dict,in_school_dict: dict, title_row_size: int, heading_three_exist: bool, class_num: int) -> dict:
    """
    筛选出指定班级的学生

    :param workbook: 指定工作簿（后缀为.xlsx的文件）
    :param sheet: 指定工作表
    :param max_row: 工作表最大值
    :param sub_title_dict: 表头位置及内容的字典
    :param in_school_dict: 学生姓名所在位置及姓名的字典
    :param title_row_size: 标题行大小
    :param heading_three_exist: 是否存在次表头
    :param class_num: 指定筛选出的班级
    :return: 返回
    """
    wb = openpyxl.load_workbook(workbook)
    sheet = wb[sheet]

    in_class_temp_dict = {}

    try:
        class_cell = [key for key, value in sub_title_dict.items() if value == "班级"][0]
        class_cell_row = ''.join(re.findall(r'\d+', class_cell))
        class_cell_col = re.match(r"^([A-Za-z]+)", class_cell).group(1).upper()

        if heading_three_exist:
            start_row = int(class_cell_row) + 1 + title_row_size
        else:
            start_row = int(class_cell_row) + title_row_size

    except IndexError:
        print("表头没有\"班级\"")

    else:
        try:
            name_cell = [key for key, value in sub_title_dict.items() if value == "姓名"][0]
            name_cell_col = re.match(r"^([A-Za-z]+)", name_cell).group(1).upper()

            for row in range(start_row, max_row + 1):
                class_place = f"{class_cell_col}{row}"
                name_place = f"{name_cell_col}{row}"
                if class_num is None:
                    in_class_temp_dict[name_place] = sheet[name_place].value
                else:
                    if class_num in sheet[class_place].value:
                        in_class_temp_dict[name_place] = sheet[name_place].value

        except IndexError:
            print("表头没有\"姓名\"")

    wb.close()

    in_class_dict = {
        key: value
        for key, value in in_school_dict.items()
        if key in in_class_temp_dict and value == in_class_temp_dict[key]
    }

    # print(f"in_class_dict内容是\n{in_class_dict}")

    return in_class_dict


def get_personal_scores(workbook: str, sheet: str, start_cell: str, max_row: int, max_column: int, title_row_size: int, heading_three_exist: bool, in_class_dict: dict) -> list:
    """
    获得学生个人成绩信息

    :param workbook: 指定工作簿（后缀为.xlsx的文件）
    :param sheet: 指定工作表
    :param start_cell: 工作表内容起始位置
    :param max_row: 工作表最大行
    :param max_column: 工作表最大列
    :param title_row_size: 工作表标题行大小
    :param heading_three_exist: 工作表是否存在次表头
    :param in_class_dict: 位置及姓名对应的字典
    :return: 有位置
    """
    wb = openpyxl.load_workbook(workbook)
    sheet = wb[sheet]

    personal_scores_list = []
    in_class_row_list = []
    start_cell_column = re.match(r"^([A-Za-z]+)", start_cell).group(1).upper()

    if heading_three_exist:
        start_record_row = title_row_size + 3
    else:
        start_record_row = title_row_size + 2

    for key in in_class_dict.keys():
        key_row = int(''.join(re.findall(r'\d+', key)))
        in_class_row_list.append(key_row)

    for row in range(start_record_row, max_row + 1):

        row_list = []

        if row in in_class_row_list:

            for column in range(column_index_from_string(start_cell_column), max_column + 1):

                row_list.append(sheet.cell(row=row, column=column).value)

            personal_scores_list.append(row_list)

    wb.close()

    # print(f"personal_scores_list中的内容为{personal_scores_list}")

    return personal_scores_list


def calc_total_average(workbook: str, sheet: str, column_list: list, mode: str) -> None:
    """
    计算指定列（科目） 的得分平均值

    :param workbook: 指定工作簿（后缀为.xlsx的文件）
    :param sheet: 指定工作表
    :param column_list: 指定的要进行计算列的列表
    :param mode: 计算模式，normal模式不去零，反之去零
    :return: 返回值为None
    """
    wb = openpyxl.load_workbook(workbook)
    sheet = wb[sheet]

    max_row = sheet.max_row
    total_score = 0
    count = 0

    try:
        if mode == "normal":

            for column in column_list:

                for row in range(1, max_row + 1):

                    cell_value = sheet[f"{column}{row}"].value

                    if isinstance(cell_value, (int, float)):
                        count += 1
                        total_score += cell_value

                    average_score = total_score / count
                    sheet[f"{column}{max_row+1}"].value = average_score

        elif mode == "normal no zero":

            for column in column_list:

                for row in range(1, max_row + 1):

                    cell_value = sheet[f"{column}{row}"].value

                    if isinstance(cell_value, (int, float)) and cell_value != 0:
                        count += 1
                        total_score += cell_value

                average_score = total_score / count
                sheet[f"{column}{max_row+1}"].value = average_score

    except:

        print("计算平均值时，列或者模式输入错误！请您重新输入！")

    else:
        print(f"已完成计算平均值，进行了{len(column_list)}计算")

    wb.save(workbook)

    wb.close()

    return None


def range_by_num(workbook: str, sheet: str, title_row_size: int, heading_three_exist: bool) -> None:
    """
    对指定工作表进行填充排序

    :param workbook: 指定工作簿（后缀为.xlsx的文件）
    :param sheet: 指定工作表
    :param title_row_size: 标题行大小
    :param heading_three_exist: 次表头是否存在
    :return: 返回值为0
    """
    wb = openpyxl.load_workbook(workbook)
    sheet = wb[sheet]

    max_row = sheet.max_row
    max_col = sheet.max_column

    if heading_three_exist:

        start_row = title_row_size + 3

        for row in range(start_row, max_row + 1):
            sheet.cell(row=row, column=max_col + 1).value = row - 3

    else:

        start_row = title_row_size + 2

        for row in range(start_row, max_row + 1):
            sheet.cell(row=row, column=max_col + 1).value = row - 2

    wb.save(workbook)

    wb.close()

    return None


def mark_subject_scores(workbook: str, sheet: str, column_list: list, color: str) -> int:
    """
    对指定列的最大值进行颜色标记

    :param workbook: 指定工作簿（后缀为.xlsx的文件）
    :param sheet: 指定工作表
    :param column_list: 需要标记的列表
    :param color: 进行标记的的颜色
    :return: 返回标记次数
    """
    wb = openpyxl.load_workbook(workbook)
    sheet = wb[sheet]

    column_max_dict = {}
    targeted_num = 0
    max_row = sheet.max_row
    max_column = sheet.max_column

    if len(column_list) == 0:
        for column in range(1, max_column + 1):
            max_num = 0
            for row in range(1, max_row + 1):
                if isinstance(sheet.cell(row=row, column=column).value, (int, float)) and sheet.cell(row=row,column=column).value > max_num:
                    max_num = sheet.cell(row=row, column=column).value
            column_max_dict[get_column_letter(column)] = max_num

        for key, value in column_max_dict.items():

            for row in range(1, max_row + 1):
                cell = sheet[f"{key}{row}"]
                cell_value = cell.value

                if isinstance(cell_value, (int, float)) and cell_value == value:
                    cell.font = Font(color=color, bold=True)
                    targeted_num += 1

    else:
        for column in column_list:

            max_num = 0

            for row in range(1, max_row + 1):

                if isinstance(sheet[f"{column}{row}"].value, (int, float)) and sheet[f"{column}{row}"].value > max_num:
                    max_num = sheet[f"{column}{row}"].value

            column_max_dict[column] = max_num

        for key, value in column_max_dict.items():

            for row in range(1, max_row + 1):
                cell = sheet[f"{key}{row}"]
                cell_value = cell.value

                if isinstance(cell_value, (int, float)) and cell_value == value:
                    cell.font = Font(color=color, bold=True)
                    targeted_num += 1

    wb.save(workbook)

    wb.close()

    return targeted_num


def compare_nums(workbook: str, sheet: str, first_column: str, second_colum: str, title_row_size: int, heading_three_exist: bool, column_name: str) -> None:
   """
   对指定的第二列减去第一列获取的新列

   :param workbook: 指定工作簿（后缀为.xlsx的文件）
   :param sheet: 指定工作表
   :param first_column: 需要进行计算的第一列，作为被减列
   :param second_colum: 作为进行计算的第二列，作为减列
   :param title_row_size: 标题的行大小
   :param heading_three_exist: 次表头是否存在
   :param column_name: 新列的名称
   :return: 返回值为None
   """
   wb = openpyxl.load_workbook(workbook)
   sheet = wb[sheet]
   # 这类对最后列或行的操作，在内部函数获得最大行、最大列，可以避免一些不必要的麻烦
   max_row = sheet.max_row
   max_column = sheet.max_column

   for row in range(1, max_row + 1):
       if isinstance(sheet[f"{first_column}{row}"].value, (int, float)):

            if isinstance(sheet[f"{second_colum}{row}"].value, (int, float)):
               compare_num = sheet[f"{second_colum}{row}"].value - sheet[f"{first_column}{row}"].value
               sheet.cell(row=row, column=max_column + 1).value = compare_num

   # 如果次表头存在或不存在，相应指定到该表最后一列对表头或次表头进行命名
   sheet.cell(row=2, column=max_column + 1).value = column_name

   if heading_three_exist:
       sheet.cell(row=3, column=max_column + 1).value = column_name

   wb.save(workbook)

   wb.close()

   return None


def string_to_num(workbook: str, sheet: str, max_row: int, max_column: int) -> None:
    """
    工作表可能出现字符串数字，将其转化为数字

    :param workbook: 指定工作簿（后缀为.xlsx的文件）
    :param sheet: 指定工作表
    :param max_row: 工作表最大行
    :param max_column: 工作表最大列
    :return: 返回值为None
    """
    wb = openpyxl.load_workbook(workbook)
    sheet = wb[sheet]

    # 定义初始值，防止表内无可从字符串数字转化到真数字的字符串
    changed_num = 0

    for column in range(1, max_column + 1):

        for row in range(1, max_row + 1):

            # 对字符串进行转换尝试
            try:
                cell_value = sheet.cell(row=row, column=column).value
                float_value = float(cell_value)

            except (ValueError, TypeError):
                continue

            else:
                changed_num += 1

                # 去掉包含0的字符串，将其转化为整型
                if float_value.is_integer():
                    sheet.cell(row=row, column=column).value = int(float_value)
                else:
                    sheet.cell(row=row, column=column).value = float_value

    wb.save(workbook)

    wb.close()

    print(f"工作表已转化{changed_num}次字符串到整型或浮点型")

    return None


def create_new_workbook(workbook: str, sheet: str, max_column: int, start_cell: str, title: str, sub_title: dict, heading_three: list, title_exist: bool,  heading_three_exist: bool, data: list) -> None:
    """
    创建新的工作簿及工作表

    :param workbook: 新命名的工作簿
    :param sheet: 新命名的工作表
    :param max_column: 原来
    :param start_cell: 原工作表最大列
    :param title: 工作表内容的标题
    :param sub_title: 原工作表的表头
    :param heading_three: 原工作表的次表头
    :param title_exist: 原内容标题是否存在
    :param heading_three_exist: 原次表头是否存在
    :param data: 携带者主要内容的字典
    :return: 返回None
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet

    # 定义新表最大列
    start_cell_column = re.match(r"^([A-Za-z]+)", start_cell).group(1).upper()
    max_column = max_column - column_index_from_string(start_cell_column) + 1

    # 对标题和次表头进行一系列判断，但是不变的是表头必定存在，有两个改动即标题只占一行且所有内容必定紧贴左上角
    if title_exist is True and heading_three_exist is False:

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_column)
        ws['A1'] = title

        for col, header in enumerate(sub_title.values(), 1):
            ws.cell(row=2, column=col, value=header)

        for row_idx, score in enumerate(data, start=3):
            for col_idx, value in enumerate(score, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

    elif title_exist is False and heading_three_exist is False:

        for col, header in enumerate(sub_title.values(), 1):
            ws.cell(row=1, column=col, value=header)

        for row_idx, score in enumerate(data, start=2):
            for col_idx, value in enumerate(score, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

    elif title_exist is True and heading_three_exist is True:

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_column)
        ws['A1'] = title

        for col, header in enumerate(sub_title.values(), 1):
            ws.cell(row=2, column=col, value=header)

        for col, header in enumerate(heading_three, 1):
            ws.cell(row=3, column=col, value=header)

        for row_idx, score in enumerate(data, start=4):
            for col_idx, value in enumerate(score, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

    elif title_exist is False and heading_three_exist is True:

        for col, header in enumerate(sub_title.values(), 1):
            ws.cell(row=1, column=col, value=header)

        for col, header in enumerate(heading_three, 1):
            ws.cell(row=2, column=col, value=header)

        for row_idx, score in enumerate(data, start=3):
            for col_idx, value in enumerate(score, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

    wb.save(workbook)

    wb.close()

    print("已成功创建新文件")

    return None


def main():
    parser = argparse.ArgumentParser(
        description="Excel学生成绩处理工具",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter
    )

    # 主要参数
    parser.add_argument("subject", type=str, help="科目名称（物理/历史）")
    parser.add_argument("filename", type=str, help="文件的另存为名称（已自动后缀.xlsx）")
    parser.add_argument("title", type=str, help="工作表内标题名称，同时作为工作表名称")
    parser.add_argument("-d", "--directory", default=".", type=str, help="在工作簿所在目录筛选出指定工作簿")
    parser.add_argument("-s", "--sheet", default=None, type=str, help="筛选出指定工作表")
    parser.add_argument("-sc", "--school", default=None, type=str, help="指定筛选出学校学生")
    parser.add_argument("-cr", "--classr", default=None, type=str, help="指定筛选出班级学生")
    parser.add_argument("-rn","--rank-number", default=False, choices=[True, False], type=bool, help="是否启用对该指定工作表进行排序，插入到最后一列")
    parser.add_argument("-mn", "--mark-column",nargs="+", type=str,  help="对指定列中最大值进行标记")
    parser.add_argument("-mr", "--mark-color", default="FF0000", type=str, help="对指定列最大值进行标记的颜色，默认颜色为红色")
    parser.add_argument("-ctac", "--calc-total-average-column" , default=None, nargs="+", type=str, help="对指定列进行计算平均值")
    parser.add_argument("-ctam", "--calc-total-average-mode", type=str, default="normal no zero", choices=["normal", "normal no zero"], help="计算一列（学科）的平均值并附加在新的最后一行")
    parser.add_argument("-cn", "--compare-nums", nargs=3, metavar=('FirstColumn', "SecondColumn", "ColumnName"), help="对比数字，由第二列减去第一列，对比后的数字放在新加最后一列\nFirstColumn： 对比的第一列\nSecondColumn：对比的第二列\nColumnName: 新列名称")

    args = parser.parse_args()

    input_workbook = get_workbook(args.subject, args.directory)
    input_sheet = get_sheet(input_workbook, args.sheet, args.subject)
    max_row, max_column, data_place = get_data_place(input_workbook, input_sheet)
    title_row_size, title_col_size, title, title_exist = verify_title(input_workbook, input_sheet, data_place)
    sub_title_place, sub_title_dict = get_sub_title(input_workbook, input_sheet, data_place, title_row_size, title_col_size, max_column)
    heading_three_list, heading_three_exist = verify_heading_three(input_workbook, input_sheet, sub_title_place, max_column)
    school_cell, in_school_student_dict = verify_school(input_workbook, input_sheet, max_row, sub_title_dict, title_row_size, heading_three_exist, args.school)
    in_class_student_dict = verify_class(input_workbook, input_sheet, max_row, sub_title_dict, in_school_student_dict, title_row_size, heading_three_exist, args.classr)
    main_content_list = get_personal_scores(input_workbook, input_sheet, data_place, max_row, max_column, title_row_size, heading_three_exist, in_class_student_dict)
    new_workbook = args.filename + ".xlsx"
    new_sheet = args.title
    safe_open_and_close(args.directory + "\\" + new_workbook)
    create_new_workbook(new_workbook, new_sheet, max_column, data_place, args.title, sub_title_dict,heading_three_list, title_exist, heading_three_exist, main_content_list)

    new_max_row, new_max_column, new_data_place = get_data_place(new_workbook, new_sheet)
    new_title_row_size, new_title_col_size, new_title, new_title_exist = verify_title(new_workbook, new_sheet, new_data_place)
    string_to_num(new_workbook, new_sheet, new_max_row, new_max_column)

    if args.compare_nums:
        try:
            first_column = str(args.compare_nums[0])
            second_column = str(args.compare_nums[1])
            column_name = str(args.compare_nums[2])

            compare_nums(new_workbook, args.title, first_column, second_column, title_row_size, heading_three_exist, column_name)

        except(ValueError, IndexError) as e:
            print(f"参数错误：{str(e)}")
            print("请使用格式：-cn <字符串> <字符串> <字符串>")

    else:
        print("已关闭对比")

    if args.rank_number:
        range_by_num(new_workbook, args.title, new_title_row_size, heading_three_exist)
        print("已启用排序，排序放置在新的最后一列")
    else:
        print("已关闭排序")

    if args.calc_total_average_column:
        try:
            calc_total_average(new_workbook, new_sheet, args.calc_total_average_column, args.calc_total_average_mode)
        except:
            print("在计算过程中，您的输入有误！请确认后重新输入")

    if args.mark_column:
        try:
            marked_times = mark_subject_scores(new_workbook, args.title, args.mark_column, args.mark_color)

        except:
            print("在标记过程中，您的输入有误！请确认后重新输入")

        else:

            print(f"工作表中有{marked_times}次标记")


if __name__ == "__main__":
    main()